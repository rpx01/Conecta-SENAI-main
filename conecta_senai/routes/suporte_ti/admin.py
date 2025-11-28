"""Rotas administrativas do módulo de suporte de TI."""
from __future__ import annotations

import csv
import io
import os
from datetime import datetime, timezone, timedelta
from typing import Iterable

from flask import Blueprint, jsonify, request, g, current_app, make_response
from sqlalchemy import func
from sqlalchemy.exc import SQLAlchemyError

from conecta_senai.auth import admin_required
from conecta_senai.models import db
from conecta_senai.models.suporte_basedados import SuporteArea, SuporteTipoEquipamento
from conecta_senai.models.suporte_chamado import SuporteChamado
from conecta_senai.routes.suporte_ti.utils import ensure_tables_exist

suporte_ti_admin_bp = Blueprint(
    "suporte_ti_admin",
    __name__,
    url_prefix="/api/suporte_ti/admin",
)


CANONICAL_STATUS = {"Aberto", "Em Atendimento", "Finalizado", "Cancelado"}

# Sinônimos aceitos para manter compatibilidade com registros antigos ou requisições
# que ainda utilizem a nomenclatura anterior dos status.
STATUS_ALIASES = {
    "Em Andamento": "Em Atendimento",
    "Fechado": "Finalizado",
}

_STATUS_ALIAS_LOOKUP = {alias.lower(): canonical for alias, canonical in STATUS_ALIASES.items()}
_CANONICAL_LOOKUP = {status.lower(): status for status in CANONICAL_STATUS}


def _normalizar_status(valor: str | None) -> str | None:
    """Converte o status informado para sua forma canônica quando aplicável."""

    if not valor:
        return valor
    valor_limpo = valor.strip()
    if not valor_limpo:
        return valor_limpo
    chave = valor_limpo.lower()
    if chave in _STATUS_ALIAS_LOOKUP:
        return _STATUS_ALIAS_LOOKUP[chave]
    if chave in _CANONICAL_LOOKUP:
        return _CANONICAL_LOOKUP[chave]
    return valor_limpo


def _obter_equivalentes_status(statuses: Iterable[str]) -> set[str]:
    """Gera o conjunto de status equivalentes (canônicos + sinônimos) para consulta."""

    equivalentes: set[str] = set()
    for status in statuses:
        if not status:
            continue
        canonico = _normalizar_status(status)
        if not canonico:
            continue
        equivalentes.add(canonico)
        for alias, destino in STATUS_ALIASES.items():
            if destino == canonico:
                equivalentes.add(alias)
    return equivalentes


def _serialize_chamado(chamado: SuporteChamado) -> dict:
    return {
        "id": chamado.id,
        "user_id": chamado.user_id,
        "nome": chamado.user.nome if chamado.user else chamado.nome_solicitante,
        "nome_solicitante": chamado.nome_solicitante,
        "email": chamado.email,
        "area": chamado.area,
        "tipo_equipamento_id": chamado.tipo_equipamento_id,
        "tipo_equipamento_nome": (
            chamado.tipo_equipamento.nome if chamado.tipo_equipamento else None
        ),
        "patrimonio": chamado.patrimonio,
        "numero_serie": chamado.numero_serie,
        "descricao_problema": chamado.descricao_problema,
        "nivel_urgencia": chamado.nivel_urgencia,
        "status": _normalizar_status(chamado.status) or chamado.status,
        "created_at": chamado.created_at.isoformat() if chamado.created_at else None,
        "updated_at": chamado.updated_at.isoformat() if chamado.updated_at else None,
        "inicio_atendimento_at": (
            chamado.inicio_atendimento_at.isoformat()
            if chamado.inicio_atendimento_at
            else None
        ),
        "encerrado_at": (
            chamado.encerrado_at.isoformat() if chamado.encerrado_at else None
        ),
        "observacoes": chamado.observacoes,
        "local_unidade": chamado.local_unidade,
        "anexos": [anexo.file_path for anexo in chamado.anexos],
    }


@suporte_ti_admin_bp.route("/todos_chamados", methods=["GET"])
@admin_required
def listar_todos_chamados():
    ensure_tables_exist([SuporteChamado])
    consulta = SuporteChamado.query

    status_param = request.args.get("status")
    if status_param:
        status_lista = [valor.strip() for valor in status_param.split(",") if valor.strip()]
        equivalentes = _obter_equivalentes_status(status_lista)
        if equivalentes:
            consulta = consulta.filter(SuporteChamado.status.in_(equivalentes))

    area_param = request.args.get("area")
    if area_param:
        consulta = consulta.filter(SuporteChamado.area == area_param)

    tipo_param = request.args.get("tipo_equipamento_id") or request.args.get(
        "tipoEquipamentoId"
    )
    if tipo_param:
        try:
            tipo_id = int(tipo_param)
            consulta = consulta.filter(SuporteChamado.tipo_equipamento_id == tipo_id)
        except ValueError:
            return jsonify({"erro": "Parâmetro tipo_equipamento_id inválido"}), 400

    urgencia_param = request.args.get("nivel_urgencia")
    if urgencia_param:
        consulta = consulta.filter(SuporteChamado.nivel_urgencia == urgencia_param)

    data_inicio = request.args.get("data_inicio") or request.args.get("dataInicio")
    data_fim = request.args.get("data_fim") or request.args.get("dataFim")

    def _parse_data(valor: str | None) -> datetime | None:
        if not valor:
            return None
        try:
            return datetime.fromisoformat(valor)
        except ValueError:
            try:
                return datetime.strptime(valor, "%Y-%m-%d")
            except ValueError:
                return None

    inicio = _parse_data(data_inicio)
    fim = _parse_data(data_fim)

    if data_inicio and inicio is None:
        return jsonify({"erro": "data_inicio inválida"}), 400
    if data_fim and fim is None:
        return jsonify({"erro": "data_fim inválida"}), 400

    if inicio is not None:
        consulta = consulta.filter(SuporteChamado.created_at >= inicio)
    if fim is not None:
        consulta = consulta.filter(SuporteChamado.created_at <= fim)

    consulta = consulta.order_by(SuporteChamado.created_at.desc())
    chamados = consulta.all()
    return jsonify([_serialize_chamado(chamado) for chamado in chamados])


@suporte_ti_admin_bp.route("/chamados/exportar_excel", methods=["GET"])
@admin_required
def exportar_chamados_excel():
    """Exporta todos os chamados do sistema em formato XLSX (Excel)."""
    current_app.logger.info("XLSX EXPORT: Iniciando exportação de chamados em formato XLSX")
    ensure_tables_exist([SuporteChamado])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        current_app.logger.info("XLSX EXPORT: openpyxl importado com sucesso")
    except ImportError:
        current_app.logger.error("XLSX EXPORT: openpyxl não está instalado")
        return jsonify({"erro": "Biblioteca openpyxl não está instalada"}), 500

    # Buscar todos os chamados sem filtro de status
    chamados = (
        SuporteChamado.query
        .order_by(SuporteChamado.created_at.asc())
        .all()
    )

    def _fmt(dt):
        """Formatar datetime para string."""
        return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""

    # Criar workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados Suporte TI"

    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Cabeçalhos
    headers = [
        "ID",
        "Nome solicitante",
        "Email",
        "Área",
        "Tipo de equipamento",
        "Nível urgência",
        "Status",
        "Abertura",
        "Início atendimento",
        "Encerramento",
        "Última atualização",
        "Observações",
    ]

    # Escrever cabeçalhos
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Escrever dados
    for row_num, c in enumerate(chamados, 2):
        ws.cell(row=row_num, column=1, value=c.id)
        ws.cell(row=row_num, column=2, value=c.nome_solicitante or (c.user.nome if c.user else ""))
        ws.cell(row=row_num, column=3, value=c.email)
        ws.cell(row=row_num, column=4, value=c.area or "")
        ws.cell(row=row_num, column=5, value=c.tipo_equipamento.nome if c.tipo_equipamento else "")
        ws.cell(row=row_num, column=6, value=c.nivel_urgencia or "")
        ws.cell(row=row_num, column=7, value=c.status or "")
        ws.cell(row=row_num, column=8, value=_fmt(c.created_at))
        ws.cell(row=row_num, column=9, value=_fmt(c.inicio_atendimento_at))
        ws.cell(row=row_num, column=10, value=_fmt(c.encerrado_at))
        ws.cell(row=row_num, column=11, value=_fmt(c.updated_at))
        ws.cell(row=row_num, column=12, value=(c.observacoes or ""))

    # Ajustar largura das colunas
    column_widths = [8, 25, 30, 20, 20, 15, 15, 20, 20, 20, 20, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    # Salvar em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Preparar resposta
    current_app.logger.info(f"XLSX EXPORT: Arquivo gerado com sucesso. Tamanho: {len(output.getvalue())} bytes")
    resposta = make_response(output.getvalue())
    resposta.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resposta.headers["Content-Disposition"] = 'attachment; filename="chamados_suporte_ti.xlsx"'
    current_app.logger.info("XLSX EXPORT: Headers configurados. Retornando arquivo XLSX")
    return resposta


@suporte_ti_admin_bp.route("/chamados/<int:chamado_id>/status", methods=["PUT"])
@admin_required
def atualizar_status_chamado(chamado_id: int):
    """Atualiza o status de um chamado existente."""

    ensure_tables_exist([SuporteChamado])

    dados = request.get_json(silent=True) or {}
    novo_status = (dados.get("status") or "").strip()
    status_normalizado = _normalizar_status(novo_status)

    if not novo_status:
        return jsonify({"erro": "Status é obrigatório."}), 400

    if not status_normalizado or status_normalizado not in CANONICAL_STATUS:
        return jsonify({"erro": "Status informado é inválido."}), 400

    chamado = db.session.get(SuporteChamado, chamado_id)
    if not chamado:
        return jsonify({"erro": "Chamado não encontrado."}), 404

    # Salvar status anterior para comparação
    status_anterior = chamado.status

    # Timezone de Brasília (UTC-3)
    tz_brasilia = timezone(timedelta(hours=-3))
    agora = datetime.now(tz_brasilia).replace(tzinfo=None)

    # Lógica de registro de timestamps baseada em transições de status
    # Normalizar status anterior também para comparação consistente
    status_anterior_normalizado = _normalizar_status(status_anterior) if status_anterior else None

    # Registrar início de atendimento quando transiciona PARA "Em Atendimento"
    # e ainda não possui esse timestamp
    if status_normalizado == "Em Atendimento" and not chamado.inicio_atendimento_at:
        chamado.inicio_atendimento_at = agora
        current_app.logger.info(
            f"Registrado inicio_atendimento_at para chamado {chamado_id}: "
            f"status {status_anterior_normalizado} → {status_normalizado}"
        )

    # Registrar encerramento quando transiciona PARA "Finalizado" ou "Cancelado"
    # e ainda não possui esse timestamp
    if status_normalizado in ("Finalizado", "Cancelado") and not chamado.encerrado_at:
        # Se não tem início de atendimento, registra agora também
        if not chamado.inicio_atendimento_at:
            chamado.inicio_atendimento_at = agora
            current_app.logger.info(
                f"Registrado inicio_atendimento_at (retroativo) para chamado {chamado_id} "
                f"pois foi encerrado sem atendimento prévio"
            )
        chamado.encerrado_at = agora
        current_app.logger.info(
            f"Registrado encerrado_at para chamado {chamado_id}: "
            f"status {status_anterior_normalizado} → {status_normalizado}"
        )

    chamado.status = status_normalizado
    chamado.updated_at = agora

    if "observacoes" in dados:
        valor_observacoes = dados.get("observacoes")
        if isinstance(valor_observacoes, str):
            valor_limpo = valor_observacoes.strip()
            chamado.observacoes = valor_limpo or None
        elif valor_observacoes is None:
            chamado.observacoes = None
        else:
            chamado.observacoes = str(valor_observacoes)

    try:
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return jsonify({"erro": "Não foi possível atualizar o status do chamado."}), 500

    return (
        jsonify(
            {
                "mensagem": "Status atualizado com sucesso.",
                "status": chamado.status,
                "updated_at": chamado.updated_at.isoformat() if chamado.updated_at else None,
            }
        ),
        200,
    )


@suporte_ti_admin_bp.route("/chamados/<int:chamado_id>", methods=["PUT"])
@admin_required
def atualizar_chamado(chamado_id: int):
    """Atualiza os dados principais de um chamado existente."""

    if not _eh_admin_raiz():
        return _resposta_nao_autorizado()

    ensure_tables_exist([SuporteChamado, SuporteArea, SuporteTipoEquipamento])

    payload = request.get_json(silent=True) or {}
    if not payload:
        return jsonify({"erro": "Corpo da requisição ausente."}), 400

    chamado = db.session.get(SuporteChamado, chamado_id)
    if not chamado:
        return jsonify({"erro": "Chamado não encontrado."}), 404

    campos_atualizados = False

    area_id = payload.get("area_id")
    area_nome = payload.get("area")
    if area_id is not None:
        try:
            area_id_int = int(area_id)
        except (TypeError, ValueError):
            return jsonify({"erro": "Identificador de área inválido."}), 400
        area_registro = db.session.get(SuporteArea, area_id_int)
        if not area_registro:
            return jsonify({"erro": "Área não encontrada."}), 404
        chamado.area = area_registro.nome
        campos_atualizados = True
    elif isinstance(area_nome, str) and area_nome.strip():
        chamado.area = area_nome.strip()
        campos_atualizados = True

    tipo_id = payload.get("tipo_equipamento_id")
    if tipo_id is None:
        tipo_id = payload.get("equipamento_id")
    if tipo_id is not None:
        try:
            tipo_id_int = int(tipo_id)
        except (TypeError, ValueError):
            return jsonify({"erro": "Identificador de equipamento inválido."}), 400
        tipo_registro = db.session.get(SuporteTipoEquipamento, tipo_id_int)
        if not tipo_registro:
            return jsonify({"erro": "Tipo de equipamento não encontrado."}), 404
        chamado.tipo_equipamento_id = tipo_registro.id
        campos_atualizados = True

    descricao = payload.get("descricao_problema")
    if descricao is None:
        descricao = payload.get("descricao")
    if descricao is None:
        descricao = payload.get("titulo")
    if isinstance(descricao, str):
        chamado.descricao_problema = descricao.strip()
        campos_atualizados = True

    urgencia = payload.get("nivel_urgencia")
    if urgencia is None:
        urgencia = payload.get("urgencia")
    urgencia_normalizada = _normalizar_urgencia(urgencia)
    if urgencia is not None:
        if not urgencia_normalizada or urgencia_normalizada not in _URGENCIAS_VALIDAS:
            return jsonify({"erro": "Nível de urgência inválido."}), 400
        chamado.nivel_urgencia = urgencia_normalizada
        campos_atualizados = True

    if "patrimonio" in payload:
        valor = payload.get("patrimonio")
        if isinstance(valor, str):
            valor = valor.strip() or None
        chamado.patrimonio = valor
        campos_atualizados = True

    if "numero_serie" in payload or "numeroSerie" in payload:
        valor = payload.get("numero_serie")
        if valor is None:
            valor = payload.get("numeroSerie")
        if isinstance(valor, str):
            valor = valor.strip() or None
        chamado.numero_serie = valor
        campos_atualizados = True

    if "observacoes" in payload:
        valor = payload.get("observacoes")
        if isinstance(valor, str):
            valor = valor.strip() or None
        chamado.observacoes = valor
        campos_atualizados = True

    if not campos_atualizados:
        return jsonify({"erro": "Nenhum campo válido para atualização foi informado."}), 400

    # Timezone de Brasília (UTC-3)
    tz_brasilia = timezone(timedelta(hours=-3))
    chamado.updated_at = datetime.now(tz_brasilia).replace(tzinfo=None)

    try:
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return jsonify({"erro": "Não foi possível atualizar o chamado."}), 500

    return (
        jsonify(
            {
                "mensagem": "Chamado atualizado com sucesso.",
                "chamado": _serialize_chamado(chamado),
            }
        ),
        200,
    )


@suporte_ti_admin_bp.route("/chamados/<int:chamado_id>", methods=["DELETE"])
@admin_required
def excluir_chamado(chamado_id: int):
    """Remove definitivamente um chamado."""

    if not _eh_admin_raiz():
        return _resposta_nao_autorizado()

    ensure_tables_exist([SuporteChamado])

    chamado = db.session.get(SuporteChamado, chamado_id)
    if not chamado:
        return jsonify({"erro": "Chamado não encontrado."}), 404

    try:
        db.session.delete(chamado)
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return jsonify({"erro": "Não foi possível excluir o chamado."}), 500

    return jsonify({"mensagem": "Chamado excluído com sucesso."}), 200


@suporte_ti_admin_bp.route("/indicadores", methods=["GET"])
@admin_required
def obter_indicadores():
    """Retorna indicadores de suporte com filtros e métricas de tempo."""
    ensure_tables_exist([SuporteChamado])

    # Construir query base
    query = db.session.query(SuporteChamado)

    # Aplicar filtros da query string
    data_inicio_str = request.args.get("data_inicio")
    data_fim_str = request.args.get("data_fim")
    area = request.args.get("area")
    tipo_equipamento_id = request.args.get("tipo_equipamento_id")
    nivel_urgencia = request.args.get("nivel_urgencia")
    status = request.args.get("status")

    # Filtro de data início
    if data_inicio_str:
        try:
            # Timezone de Brasília (UTC-3)
            tz_brasilia = timezone(timedelta(hours=-3))
            data_inicio = datetime.fromisoformat(data_inicio_str).replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=tz_brasilia).replace(tzinfo=None)
            query = query.filter(SuporteChamado.created_at >= data_inicio)
        except (ValueError, TypeError):
            pass

    # Filtro de data fim
    if data_fim_str:
        try:
            tz_brasilia = timezone(timedelta(hours=-3))
            data_fim = datetime.fromisoformat(data_fim_str).replace(hour=23, minute=59, second=59, microsecond=999999, tzinfo=tz_brasilia).replace(tzinfo=None)
            query = query.filter(SuporteChamado.created_at <= data_fim)
        except (ValueError, TypeError):
            pass

    # Filtro de área
    if area:
        query = query.filter(SuporteChamado.area == area)

    # Filtro de tipo de equipamento
    if tipo_equipamento_id:
        try:
            tipo_id = int(tipo_equipamento_id)
            query = query.filter(SuporteChamado.tipo_equipamento_id == tipo_id)
        except (ValueError, TypeError):
            pass

    # Filtro de nível de urgência
    if nivel_urgencia:
        query = query.filter(SuporteChamado.nivel_urgencia == nivel_urgencia)

    # Filtro de status
    if status:
        query = query.filter(SuporteChamado.status == status)

    # Estatísticas básicas (com filtros aplicados)
    total = query.count()
    
    por_status = (
        db.session.query(SuporteChamado.status, func.count(SuporteChamado.id))
        .filter(SuporteChamado.id.in_(query.with_entities(SuporteChamado.id).subquery()))
        .group_by(SuporteChamado.status)
        .all()
    )
    
    por_tipo = (
        db.session.query(SuporteTipoEquipamento.nome, func.count(SuporteChamado.id))
        .join(
            SuporteTipoEquipamento,
            SuporteTipoEquipamento.id == SuporteChamado.tipo_equipamento_id,
            isouter=True,
        )
        .filter(SuporteChamado.id.in_(query.with_entities(SuporteChamado.id).subquery()))
        .group_by(SuporteTipoEquipamento.nome)
        .all()
    )
    
    por_urgencia = (
        db.session.query(SuporteChamado.nivel_urgencia, func.count(SuporteChamado.id))
        .filter(SuporteChamado.id.in_(query.with_entities(SuporteChamado.id).subquery()))
        .group_by(SuporteChamado.nivel_urgencia)
        .all()
    )

    # Métricas de tempo - Tempo médio até início de atendimento (em segundos)
    tempo_medio_atendimento = (
        db.session.query(
            func.avg(
                func.extract('epoch', SuporteChamado.inicio_atendimento_at) - 
                func.extract('epoch', SuporteChamado.created_at)
            )
        )
        .filter(SuporteChamado.inicio_atendimento_at.isnot(None))
        .filter(SuporteChamado.id.in_(query.with_entities(SuporteChamado.id).subquery()))
        .scalar()
    )

    # Tempo médio até encerramento (em segundos)
    tempo_medio_encerramento = (
        db.session.query(
            func.avg(
                func.extract('epoch', SuporteChamado.encerrado_at) - 
                func.extract('epoch', SuporteChamado.created_at)
            )
        )
        .filter(SuporteChamado.encerrado_at.isnot(None))
        .filter(SuporteChamado.id.in_(query.with_entities(SuporteChamado.id).subquery()))
        .scalar()
    )

    # Percentual de chamados atendidos em menos de 24 horas
    total_com_atendimento = (
        query.filter(SuporteChamado.inicio_atendimento_at.isnot(None)).count()
    )
    
    if total_com_atendimento > 0:
        atendidos_em_24h = (
            query.filter(SuporteChamado.inicio_atendimento_at.isnot(None))
            .filter(
                func.extract('epoch', SuporteChamado.inicio_atendimento_at) - 
                func.extract('epoch', SuporteChamado.created_at) < 86400
            )
            .count()
        )
        percentual_24h = (atendidos_em_24h / total_com_atendimento) * 100
    else:
        percentual_24h = 0

    # Tempo médio por urgência
    tempo_por_urgencia = []
    for nivel in ["Baixo", "Médio", "Alto"]:
        query_nivel = query.filter(SuporteChamado.nivel_urgencia == nivel)
        
        tempo_atend = (
            db.session.query(
                func.avg(
                    func.extract('epoch', SuporteChamado.inicio_atendimento_at) - 
                    func.extract('epoch', SuporteChamado.created_at)
                )
            )
            .filter(SuporteChamado.inicio_atendimento_at.isnot(None))
            .filter(SuporteChamado.id.in_(query_nivel.with_entities(SuporteChamado.id).subquery()))
            .scalar()
        )
        
        tempo_encer = (
            db.session.query(
                func.avg(
                    func.extract('epoch', SuporteChamado.encerrado_at) - 
                    func.extract('epoch', SuporteChamado.created_at)
                )
            )
            .filter(SuporteChamado.encerrado_at.isnot(None))
            .filter(SuporteChamado.id.in_(query_nivel.with_entities(SuporteChamado.id).subquery()))
            .scalar()
        )
        
        tempo_por_urgencia.append({
            "nivel": nivel,
            "tempo_atendimento": float(tempo_atend) if tempo_atend else 0,
            "tempo_encerramento": float(tempo_encer) if tempo_encer else 0,
        })

    return jsonify(
        {
            "total_chamados": total,
            "por_status": [
                {"status": status or "Não informado", "quantidade": quantidade}
                for status, quantidade in por_status
            ],
            "por_tipo_equipamento": [
                {"tipo": tipo or "Não informado", "quantidade": quantidade}
                for tipo, quantidade in por_tipo
            ],
            "por_nivel_urgencia": [
                {"nivel": nivel or "Não informado", "quantidade": quantidade}
                for nivel, quantidade in por_urgencia
            ],
            "tempo_medio_abertura_para_atendimento_segundos": float(tempo_medio_atendimento) if tempo_medio_atendimento else 0,
            "tempo_medio_abertura_para_encerramento_segundos": float(tempo_medio_encerramento) if tempo_medio_encerramento else 0,
            "percentual_atendidos_em_24h": round(percentual_24h, 2),
            "tempo_medio_por_urgencia": tempo_por_urgencia,
        }
    )


def _criar_registro_basico(model, nome: str):
    if not nome.strip():
        return None, "Nome é obrigatório."
    existente = model.query.filter(func.lower(model.nome) == nome.strip().lower()).first()
    if existente:
        return None, "Registro já cadastrado."
    registro = model(nome=nome.strip())
    try:
        db.session.add(registro)
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return None, "Erro ao salvar registro."
    return registro, None


def _atualizar_registro_basico(model, registro_id: int, nome: str):
    registro = db.session.get(model, registro_id)
    if not registro:
        return None, "Registro não encontrado."
    if not nome.strip():
        return None, "Nome é obrigatório."
    conflito = (
        model.query.filter(func.lower(model.nome) == nome.strip().lower(), model.id != registro_id)
        .first()
    )
    if conflito:
        return None, "Registro já cadastrado."
    registro.nome = nome.strip()
    try:
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return None, "Erro ao atualizar registro."
    return registro, None


def _excluir_registro_basico(model, registro_id: int):
    registro = db.session.get(model, registro_id)
    if not registro:
        return False, "Registro não encontrado."
    try:
        db.session.delete(registro)
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return False, "Erro ao excluir registro."
    return True, None


@suporte_ti_admin_bp.route("/tipos_equipamento", methods=["GET"])
@admin_required
def listar_tipos_equipamento():
    ensure_tables_exist([SuporteTipoEquipamento])
    tipos = SuporteTipoEquipamento.query.order_by(SuporteTipoEquipamento.nome.asc()).all()
    return jsonify([{"id": tipo.id, "nome": tipo.nome} for tipo in tipos])


@suporte_ti_admin_bp.route("/tipos_equipamento", methods=["POST"])
@admin_required
def criar_tipo_equipamento():
    ensure_tables_exist([SuporteTipoEquipamento])
    payload = request.get_json(silent=True) or {}
    nome = payload.get("nome", "")
    registro, erro = _criar_registro_basico(SuporteTipoEquipamento, nome)
    if erro:
        return jsonify({"erro": erro}), 400 if "salvar" not in erro.lower() else 500
    return jsonify({"id": registro.id, "nome": registro.nome}), 201


@suporte_ti_admin_bp.route("/tipos_equipamento/<int:registro_id>", methods=["PUT"])
@admin_required
def atualizar_tipo_equipamento(registro_id: int):
    ensure_tables_exist([SuporteTipoEquipamento])
    payload = request.get_json(silent=True) or {}
    nome = payload.get("nome", "")
    registro, erro = _atualizar_registro_basico(SuporteTipoEquipamento, registro_id, nome)
    if erro:
        status = 404 if "não encontrado" in erro.lower() else 400
        if "atualizar" in erro.lower():
            status = 500
        return jsonify({"erro": erro}), status
    return jsonify({"id": registro.id, "nome": registro.nome})


@suporte_ti_admin_bp.route("/tipos_equipamento/<int:registro_id>", methods=["DELETE"])
@admin_required
def excluir_tipo_equipamento(registro_id: int):
    ensure_tables_exist([SuporteTipoEquipamento])
    sucesso, erro = _excluir_registro_basico(SuporteTipoEquipamento, registro_id)
    if erro:
        status = 404 if "não encontrado" in erro.lower() else 500
        return jsonify({"erro": erro}), status
    return jsonify({"mensagem": "Tipo de equipamento removido com sucesso."})


@suporte_ti_admin_bp.route("/areas", methods=["GET"])
@admin_required
def listar_areas():
    ensure_tables_exist([SuporteArea])
    areas = SuporteArea.query.order_by(SuporteArea.nome.asc()).all()
    return jsonify([{"id": area.id, "nome": area.nome} for area in areas])


@suporte_ti_admin_bp.route("/areas", methods=["POST"])
@admin_required
def criar_area():
    ensure_tables_exist([SuporteArea])
    payload = request.get_json(silent=True) or {}
    nome = payload.get("nome", "")
    registro, erro = _criar_registro_basico(SuporteArea, nome)
    if erro:
        return jsonify({"erro": erro}), 400 if "salvar" not in erro.lower() else 500
    return jsonify({"id": registro.id, "nome": registro.nome}), 201


@suporte_ti_admin_bp.route("/areas/<int:registro_id>", methods=["PUT"])
@admin_required
def atualizar_area(registro_id: int):
    ensure_tables_exist([SuporteArea])
    payload = request.get_json(silent=True) or {}
    nome = payload.get("nome", "")
    registro, erro = _atualizar_registro_basico(SuporteArea, registro_id, nome)
    if erro:
        status = 404 if "não encontrado" in erro.lower() else 400
        if "atualizar" in erro.lower():
            status = 500
        return jsonify({"erro": erro}), status
    return jsonify({"id": registro.id, "nome": registro.nome})


@suporte_ti_admin_bp.route("/areas/<int:registro_id>", methods=["DELETE"])
@admin_required
def excluir_area(registro_id: int):
    ensure_tables_exist([SuporteArea])
    sucesso, erro = _excluir_registro_basico(SuporteArea, registro_id)
    if erro:
        status = 404 if "não encontrado" in erro.lower() else 500
        return jsonify({"erro": erro}), status
    return jsonify({"mensagem": "Área removida com sucesso."})
_URGENCIAS_VALIDAS = {"Baixo", "Médio", "Alto"}


def _eh_admin_raiz() -> bool:
    usuario = getattr(g, "current_user", None)
    if not usuario or not getattr(usuario, "email", None):
        return False
    admin_email = (os.getenv("ADMIN_EMAIL") or "").strip().lower()
    if not admin_email:
        return False
    return (usuario.email or "").strip().lower() == admin_email


def _resposta_nao_autorizado():
    return jsonify({"erro": "Apenas o Administrador raiz pode realizar esta ação."}), 403


def _normalizar_urgencia(valor: str | None) -> str | None:
    if not valor:
        return None
    valor_limpo = valor.strip()
    if not valor_limpo:
        return None
    if valor_limpo.lower() == "medio":
        valor_limpo = "Médio"
    return valor_limpo

