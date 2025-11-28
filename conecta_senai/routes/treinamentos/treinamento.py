# flake8: noqa
"""Rotas para gerenciamento de treinamentos e inscricoes."""

from flask import Blueprint, request, jsonify, g, current_app
from sqlalchemy.exc import SQLAlchemyError
import math
from datetime import date, datetime, timedelta
import logging
from types import SimpleNamespace
from pathlib import Path

from conecta_senai.models import (
    db,
    Treinamento,
    TurmaTreinamento,
    InscricaoTreinamento,
    AuditLog,
    user,
)
from conecta_senai.models.user import User
from conecta_senai.models.instrutor import Instrutor
from conecta_senai.utils.error_handler import handle_internal_error
from conecta_senai.schemas.treinamento import (
    InscricaoTreinamentoCreateSchema,
    TreinamentoCreateSchema,
    TreinamentoUpdateSchema,
    TurmaTreinamentoCreateSchema,
    TurmaTreinamentoUpdateSchema,
)
from conecta_senai.auth import login_required, admin_required
from conecta_senai.utils.audit import log_action
from pydantic import ValidationError
from io import StringIO, BytesIO
import csv
from flask import send_file, make_response
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image as ReportlabImage,
    KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from conecta_senai.services.email_service import (
    enviar_convocacao,
    send_email,
    notificar_nova_turma,
    notificar_atualizacao_turma,
    build_turma_context,
    listar_emails_secretaria,
    send_treinamento_desmarcado_email,
    send_turma_alterada_email,
)

log = logging.getLogger(__name__)

treinamento_bp = Blueprint("treinamento", __name__)


def _get_logo_path() -> Path:
    """Retorna o caminho absoluto do logo do SENAI dentro da pasta estática."""
    static_folder = current_app.static_folder
    if not static_folder:
        static_folder = str(Path(current_app.root_path) / "static")
    return Path(static_folder) / "img" / "senai-logo.png"


# Função auxiliar para coletar dados da turma
def coletar_dados_turma(turma: TurmaTreinamento) -> dict:
    """Coleta e formata os dados de um objeto TurmaTreinamento em um dicionário."""
    periodo = ""
    if turma.data_inicio and turma.data_fim:
        periodo = f"{turma.data_inicio.strftime('%d/%m/%Y')} a {turma.data_fim.strftime('%d/%m/%Y')}"
    treinamento = turma.treinamento
    return {
        "treinamento_nome": treinamento.nome if treinamento else "",
        "treinamento_codigo": treinamento.codigo if treinamento else "",
        "periodo": periodo,
        "horario": turma.horario,
        "carga_horaria": treinamento.carga_horaria if treinamento else None,
        "instrutor_nome": turma.instrutor.nome if turma.instrutor else "Não definido",
        "local_realizacao": turma.local_realizacao,
        "teoria_online": turma.teoria_online,
        "tem_pratica": treinamento.tem_pratica if treinamento else False,
        "local_pratica": getattr(turma, "local_pratica", None),
    }


# NOVO: Endpoint para listar turmas futuras (agendadas)
@treinamento_bp.route("/treinamentos/agendadas", methods=["GET"])
@login_required
def listar_turmas_agendadas():
    """Lista as turmas de treinamento que ainda não começaram."""
    hoje = date.today()
    turmas = (
        TurmaTreinamento.query.filter(TurmaTreinamento.data_inicio > hoje)
        .join(Treinamento)
        .order_by(TurmaTreinamento.data_inicio)
        .all()
    )

    dados = []
    for turma in turmas:
        dados.append(
            {
                "turma_id": turma.id,
                "treinamento": turma.treinamento.to_dict(),
                "data_inicio": (
                    turma.data_inicio.isoformat() if turma.data_inicio else None
                ),
                "data_fim": turma.data_fim.isoformat() if turma.data_fim else None,
                "local_realizacao": turma.local_realizacao,
                "horario": turma.horario,
                "instrutor_nome": turma.instrutor.nome if turma.instrutor else "A definir",
                "teoria_online": turma.teoria_online,
                "has_pratica": bool(getattr(turma.treinamento, "tem_pratica", False)),
            }
        )
    return jsonify(dados)


# NOVO: Endpoint para as turmas ativas (em andamento)
@treinamento_bp.route("/treinamentos/turmas-ativas", methods=["GET"])
@login_required
def listar_turmas_ativas():
    """Lista as turmas de treinamento que estão atualmente em andamento."""
    hoje = date.today()
    turmas = (
        TurmaTreinamento.query.filter(
            db.and_(
                TurmaTreinamento.data_inicio <= hoje, TurmaTreinamento.data_fim >= hoje
            )
        )
        .join(Treinamento)
        .order_by(TurmaTreinamento.data_inicio.desc())
        .all()
    )

    dados = []
    for turma in turmas:
        dados.append(
            {
                "turma_id": turma.id,
                "treinamento": turma.treinamento.to_dict(),
                "data_inicio": (
                    turma.data_inicio.isoformat() if turma.data_inicio else None
                ),
                "data_fim": turma.data_fim.isoformat() if turma.data_fim else None,
                "local_realizacao": turma.local_realizacao,
                "horario": turma.horario,
                "instrutor": turma.instrutor.to_dict() if turma.instrutor else None,
                "teoria_online": turma.teoria_online,
                "has_pratica": bool(getattr(turma.treinamento, "tem_pratica", False)),
            }
        )
    return jsonify(dados)


# MODIFICADO: Endpoint para o histórico de turmas (concluídas)
@treinamento_bp.route("/treinamentos/historico", methods=["GET"])
@login_required
def listar_historico_turmas():
    """Lista as turmas de treinamento que já foram concluídas."""
    hoje = date.today()
    turmas = (
        TurmaTreinamento.query.filter(TurmaTreinamento.data_fim < hoje)
        .join(Treinamento)
        .order_by(TurmaTreinamento.data_inicio.desc())
        .all()
    )

    dados = []
    for turma in turmas:
        dados.append(
            {
                "turma_id": turma.id,
                "treinamento": turma.treinamento.to_dict(),
                "data_inicio": (
                    turma.data_inicio.isoformat() if turma.data_inicio else None
                ),
                "data_fim": turma.data_fim.isoformat() if turma.data_fim else None,
                "local_realizacao": turma.local_realizacao,
                "horario": turma.horario,
                "instrutor": turma.instrutor.to_dict() if turma.instrutor else None,
                "teoria_online": turma.teoria_online,
                "has_pratica": bool(getattr(turma.treinamento, "tem_pratica", False)),
            }
        )
    return jsonify(dados)


# Se precisar retornar todas as turmas, mantém a rota original com novo nome
@treinamento_bp.route("/treinamentos/todas", methods=["GET"])
@login_required
def listar_todas_as_turmas():
    """Lista TODAS as turmas de treinamento (futuras, presentes e passadas)."""
    turmas = TurmaTreinamento.query.join(Treinamento).order_by(Treinamento.nome).all()
    dados = []
    for turma in turmas:
        dados.append(
            {
                "turma_id": turma.id,
                "treinamento": turma.treinamento.to_dict(),
                "data_inicio": (
                    turma.data_inicio.isoformat() if turma.data_inicio else None
                ),
                "data_fim": turma.data_fim.isoformat() if turma.data_fim else None,
                "local_realizacao": turma.local_realizacao,
                "horario": turma.horario,
                "instrutor": turma.instrutor.to_dict() if turma.instrutor else None,
                "teoria_online": turma.teoria_online,
                "has_pratica": bool(getattr(turma.treinamento, "tem_pratica", False)),
            }
        )
    return jsonify(dados)


@treinamento_bp.route("/treinamentos/<int:turma_id>/inscricoes", methods=["POST"])
@login_required
def inscrever_usuario(turma_id):
    """Realiza a inscricao do usuario logado em uma turma."""
    turma = db.session.get(TurmaTreinamento, turma_id)
    if not turma:
        return jsonify({"erro": "Turma não encontrada"}), 404

    existente = InscricaoTreinamento.query.filter_by(
        usuario_id=g.current_user.id, turma_id=turma_id
    ).first()
    if existente:
        return jsonify({"erro": "Usuário já inscrito nesta turma"}), 400

    data = request.json or {}
    try:
        payload = InscricaoTreinamentoCreateSchema(**data)
    except Exception as e:
        return jsonify({"erro": str(e)}), 400

    # Preenche dados a partir do perfil do usuário, se disponíveis
    usuario = g.current_user
    cpf = usuario.cpf or payload.cpf
    data_nascimento = usuario.data_nascimento or payload.data_nascimento
    empresa = usuario.empresa or payload.empresa

    try:
        insc = InscricaoTreinamento(
            usuario_id=usuario.id,
            turma_id=turma_id,
            nome=payload.nome,
            email=payload.email,
            cpf=cpf,
            data_nascimento=data_nascimento,
            empresa=empresa,
        )
        db.session.add(insc)
        db.session.commit()
        log_action(g.current_user.id, 'create', 'InscricaoTreinamento', insc.id, insc.to_dict())
        return jsonify(insc.to_dict()), 201
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route("/treinamentos/minhas", methods=["GET"])
@login_required
def listar_meus_cursos():
    """Lista cursos em que o usuario esta inscrito."""
    # Usando joinedload para otimizar a busca e evitar múltiplas queries (problema N+1)
    inscricoes = (
        InscricaoTreinamento.query.filter_by(usuario_id=g.current_user.id)
        .join(TurmaTreinamento)
        .options(
            db.joinedload(InscricaoTreinamento.turma).joinedload(
                TurmaTreinamento.treinamento
            )
        )
        .options(
            db.joinedload(InscricaoTreinamento.turma).joinedload(TurmaTreinamento.instrutor)
        )
        .all()
    )

    result = []
    for inc in inscricoes:
        turma = inc.turma
        result.append(
            {
                "id": inc.id,
                "turma_id": turma.id,
                "treinamento": turma.treinamento.to_dict(),
                "data_inicio": turma.data_inicio.isoformat() if turma.data_inicio else None,
                "data_fim": turma.data_fim.isoformat() if turma.data_fim else None,
                # Campos que estavam faltando e foram adicionados
                "horario": turma.horario,
                "local_realizacao": turma.local_realizacao,
                "instrutor_nome": turma.instrutor.nome if turma.instrutor else None,
                "teoria_online": turma.teoria_online,
                "has_pratica": bool(getattr(turma.treinamento, "tem_pratica", False)),
            }
        )
    return jsonify(result)


@treinamento_bp.route("/treinamentos/catalogo", methods=["GET"])
@login_required
def listar_catalogo_treinamentos():
    """Lista os treinamentos cadastrados."""
    treins = Treinamento.query.order_by(Treinamento.nome).all()
    return jsonify([t.to_dict() for t in treins])
@treinamento_bp.route("/treinamentos/catalogo", methods=["POST"])
@admin_required
def criar_treinamento():
    """Cadastra um novo treinamento."""
    data = request.json or {}
    try:
        payload = TreinamentoCreateSchema(**data)
    except ValidationError as e:
        return jsonify({"erro": e.errors()}), 400

    # Verificação de duplicidade de código
    if Treinamento.query.filter_by(codigo=payload.codigo).first():
        return jsonify({"erro": "Já existe um treinamento com este código"}), 400
    try:
        novo = Treinamento(
            nome=payload.nome,
            codigo=payload.codigo,
            capacidade_maxima=payload.capacidade_maxima,
            carga_horaria=payload.carga_horaria,
            tem_pratica=payload.tem_pratica,
            links_materiais=payload.links_materiais,
            tipo=payload.tipo,
            conteudo_programatico=payload.conteudo_programatico,
        )
        db.session.add(novo)
        db.session.commit()
        log_action(g.current_user.id, 'create', 'Treinamento', novo.id, novo.to_dict())
        return jsonify(novo.to_dict()), 201
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route("/treinamentos/catalogo/<int:treinamento_id>", methods=["GET"])
@login_required
def obter_treinamento(treinamento_id):
    """Obtém um treinamento específico."""
    treino = db.session.get(Treinamento, treinamento_id)
    if not treino:
        return jsonify({"erro": "Treinamento não encontrado"}), 404
    return jsonify(treino.to_dict())


@treinamento_bp.route("/treinamentos/catalogo/<int:treinamento_id>", methods=["PUT"])
@admin_required
def atualizar_treinamento(treinamento_id):
    """Atualiza um treinamento existente."""
    treino = db.session.get(Treinamento, treinamento_id)
    if not treino:
        return jsonify({"erro": "Treinamento não encontrado"}), 404
    data = request.json or {}
    try:
        payload = TreinamentoUpdateSchema(**data)
    except ValidationError as e:
        return jsonify({"erro": e.errors()}), 400

    if payload.nome is not None:
        treino.nome = payload.nome
    if payload.codigo is not None:
        existente_codigo = Treinamento.query.filter_by(codigo=payload.codigo).first()
        if existente_codigo and existente_codigo.id != treinamento_id:
            return jsonify({"erro": "Já existe um treinamento com este código"}), 400
        treino.codigo = payload.codigo
    if payload.capacidade_maxima is not None:
        treino.capacidade_maxima = payload.capacidade_maxima
    if payload.carga_horaria is not None:
        treino.carga_horaria = payload.carga_horaria
    if payload.tem_pratica is not None:
        treino.tem_pratica = payload.tem_pratica
    if payload.links_materiais is not None:
        treino.links_materiais = payload.links_materiais
    if payload.tipo is not None:
        treino.tipo = payload.tipo
    if payload.conteudo_programatico is not None:
        treino.conteudo_programatico = payload.conteudo_programatico

    try:
        db.session.commit()
        log_action(g.current_user.id, 'update', 'Treinamento', treino.id, payload.model_dump(exclude_unset=True))
        return jsonify(treino.to_dict())
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route("/treinamentos/catalogo/<int:treinamento_id>", methods=["DELETE"])
@admin_required
def remover_treinamento(treinamento_id):
    """Exclui um treinamento."""
    treino = db.session.get(Treinamento, treinamento_id)
    if not treino:
        return jsonify({"erro": "Treinamento não encontrado"}), 404
    try:
        db.session.delete(treino)
        db.session.commit()
        return jsonify({"mensagem": "Treinamento removido com sucesso"})
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route("/treinamentos/turmas", methods=["POST"])
@admin_required
def criar_turma_treinamento():
    """Cria uma turma para um treinamento."""
    data = request.json or {}
    try:
        payload = TurmaTreinamentoCreateSchema(**data)
    except ValidationError as e:
        return jsonify({"erro": e.errors()}), 400

    treinamento = db.session.get(Treinamento, payload.treinamento_id)
    if not treinamento:
        return jsonify({"erro": "Treinamento não encontrado"}), 404

    if treinamento.carga_horaria and treinamento.carga_horaria > 0:
        dias_minimos = math.ceil(treinamento.carga_horaria / 8)
        data_fim_minima = payload.data_inicio + timedelta(days=dias_minimos - 1)
        if payload.data_fim < data_fim_minima:
            return (
                jsonify(
                    {
                        "erro": f"Data de término inválida. Com base na carga horária, a data mínima é {data_fim_minima.strftime('%d/%m/%Y')}.",
                    }
                ),
                400,
            )
    turma = TurmaTreinamento(
        treinamento_id=payload.treinamento_id,
        data_inicio=payload.data_inicio,
       data_fim=payload.data_fim,
        local_realizacao=payload.local_realizacao,
        horario=payload.horario,
        instrutor_id=payload.instrutor_id,
        teoria_online=payload.teoria_online,
    )
    try:
        db.session.add(turma)
        db.session.commit()
        try:
            notificar_nova_turma(turma)
        except Exception as exc:  # pragma: no cover - log apenas
            log.error(f"Erro ao notificar nova turma: {exc}")
        return jsonify(turma.to_dict()), 201
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route("/treinamentos/turmas/<int:turma_id>", methods=["PUT"])
@admin_required
def atualizar_turma_treinamento(turma_id):
    """Atualiza uma turma."""
    turma = db.session.get(TurmaTreinamento, turma_id)
    if not turma:
        return jsonify({"erro": "Turma não encontrada"}), 404

    data_fim_turma = (
        turma.data_fim.date()
        if isinstance(turma.data_fim, datetime)
        else turma.data_fim
    )
    if data_fim_turma < date.today():
        return (
            jsonify(
                {"erro": "Não é possível modificar uma turma que já foi concluída."}
            ),
            403,
        )

    # Captura o estado ANTES de qualquer modificação
    periodo_antigo = ""
    if turma.data_inicio and turma.data_fim:
        periodo_antigo = (
            f"De {turma.data_inicio.strftime('%d/%m/%Y')} a {turma.data_fim.strftime('%d/%m/%Y')}"
        )
    dados_antigos = {
        "treinamento_nome": turma.treinamento.nome if turma.treinamento else "",
        "treinamento_codigo": turma.treinamento.codigo if turma.treinamento else "",
        "periodo": periodo_antigo,
        "horario": turma.horario,
        "carga_horaria": turma.treinamento.carga_horaria if turma.treinamento else None,
        "instrutor_nome": turma.instrutor.nome if turma.instrutor else "A definir",
        "local_realizacao": turma.local_realizacao,
        "teoria_online": turma.teoria_online,
        "tem_pratica": turma.treinamento.tem_pratica if turma.treinamento else False,
        "local_pratica": (
            getattr(turma.treinamento, "local_pratica", None)
            if turma.treinamento and turma.treinamento.tem_pratica
            else None
        ),
    }

    data = request.json or {}
    try:
        payload = TurmaTreinamentoUpdateSchema(**data)
    except ValidationError as e:
        return jsonify({"erro": e.errors()}), 400

    valores_antes = {
        "data_inicio": turma.data_inicio,
        "data_fim": turma.data_fim,
        "local_realizacao": turma.local_realizacao,
        "horario": turma.horario,
        "teoria_online": turma.teoria_online,
    }
    instrutor_antigo = turma.instrutor

    treinamento_id = (
        payload.treinamento_id if payload.treinamento_id is not None else turma.treinamento_id
    )
    treinamento = db.session.get(Treinamento, treinamento_id)
    if not treinamento:
        return jsonify({"erro": "Treinamento não encontrado"}), 404

    data_inicio = payload.data_inicio if payload.data_inicio is not None else turma.data_inicio
    data_fim = payload.data_fim if payload.data_fim is not None else turma.data_fim

    if treinamento.carga_horaria and treinamento.carga_horaria > 0:
        dias_minimos = math.ceil(treinamento.carga_horaria / 8)
        data_fim_minima = data_inicio + timedelta(days=dias_minimos - 1)
        if data_fim < data_fim_minima:
            return (
                jsonify(
                    {
                        "erro": f"Data de término inválida. Com base na carga horária, a data mínima é {data_fim_minima.strftime('%d/%m/%Y')}.",
                    }
                ),
                400,
            )

    turma.treinamento_id = treinamento_id
    turma.data_inicio = data_inicio
    turma.data_fim = data_fim
    if payload.local_realizacao is not None:
        turma.local_realizacao = payload.local_realizacao
    if payload.horario is not None:
        turma.horario = payload.horario
    if payload.instrutor_id is not None:
        if payload.instrutor_id:
            if not db.session.get(Instrutor, payload.instrutor_id):
                return jsonify({"erro": "Instrutor não encontrado"}), 404
        turma.instrutor_id = payload.instrutor_id
    if payload.teoria_online is not None:
        turma.teoria_online = payload.teoria_online
    try:
        db.session.commit()
        db.session.refresh(turma)
        periodo_novo = ""
        if turma.data_inicio and turma.data_fim:
            periodo_novo = (
                f"De {turma.data_inicio.strftime('%d/%m/%Y')} a {turma.data_fim.strftime('%d/%m/%Y')}"
            )
        dados_novos = {
            "treinamento_nome": turma.treinamento.nome if turma.treinamento else "",
            "treinamento_codigo": turma.treinamento.codigo if turma.treinamento else "",
            "periodo": periodo_novo,
            "horario": turma.horario,
            "carga_horaria": turma.treinamento.carga_horaria if turma.treinamento else None,
            "instrutor_nome": turma.instrutor.nome if turma.instrutor else "A definir",
            "local_realizacao": turma.local_realizacao,
            "teoria_online": turma.teoria_online,
            "tem_pratica": turma.treinamento.tem_pratica if turma.treinamento else False,
            "local_pratica": (
                getattr(turma.treinamento, "local_pratica", None)
                if turma.treinamento and turma.treinamento.tem_pratica
                else None
            ),
        }
        diff = {}
        fmt = "%d/%m/%Y"
        if valores_antes["data_inicio"] != turma.data_inicio:
            diff["data_inicio"] = (
                valores_antes["data_inicio"].strftime(fmt)
                if valores_antes["data_inicio"]
                else None,
                turma.data_inicio.strftime(fmt) if turma.data_inicio else None,
            )
        if valores_antes["data_fim"] != turma.data_fim:
            diff["data_fim"] = (
                valores_antes["data_fim"].strftime(fmt)
                if valores_antes["data_fim"]
                else None,
                turma.data_fim.strftime(fmt) if turma.data_fim else None,
            )
        if valores_antes["local_realizacao"] != turma.local_realizacao:
            diff["local_realizacao"] = (
                valores_antes["local_realizacao"],
                turma.local_realizacao,
            )
        if valores_antes["horario"] != turma.horario:
            diff["horario"] = (valores_antes["horario"], turma.horario)
        if valores_antes["teoria_online"] != turma.teoria_online:
            diff["teoria_online"] = (
                valores_antes["teoria_online"],
                turma.teoria_online,
            )
        if instrutor_antigo != turma.instrutor:
            diff["instrutor"] = (
                instrutor_antigo.nome if instrutor_antigo else None,
                turma.instrutor.nome if turma.instrutor else None,
            )
        if diff:
            try:
                send_turma_alterada_email(dados_antigos, dados_novos)
            except Exception as e:  # pragma: no cover - log apenas
                current_app.logger.error(
                    f"Erro ao enfileirar e-mail de alteração para turma {turma_id}: {e}"
                )
            try:
                notificar_atualizacao_turma(
                    turma, diff, instrutor_antigo, notificar_secretaria=False
                )
            except Exception as exc:  # pragma: no cover - log apenas
                log.error(f"Erro ao notificar atualização de turma: {exc}")
        log_action(
            g.current_user.id,
            'update',
            'TurmaTreinamento',
            turma.id,
            payload.model_dump(exclude_unset=True)
        )
        return jsonify(turma.to_dict())
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route("/treinamentos/turmas/<int:turma_id>", methods=["DELETE"])
@admin_required
def remover_turma_treinamento(turma_id):
    """Remove uma turma."""
    turma = db.session.get(TurmaTreinamento, turma_id)
    if not turma:
        return jsonify({"erro": "Turma não encontrada"}), 404

    data_inicio_turma = (
        turma.data_inicio.date()
        if isinstance(turma.data_inicio, datetime)
        else turma.data_inicio
    )
    if data_inicio_turma <= date.today():
        return (
            jsonify(
                {
                    "erro": "Não é possível excluir uma turma que já iniciou ou foi concluída."
                }
            ),
            403,
        )
    try:
        emails_secretaria = listar_emails_secretaria()
        instrutor_email = (
            turma.instrutor.email
            if turma.instrutor and getattr(turma.instrutor, "email", None)
            else None
        )
        recipients = []
        if emails_secretaria:
            recipients.extend(emails_secretaria)
        if instrutor_email:
            recipients.append(instrutor_email)
        unique_recipients = list(dict.fromkeys(recipients))

        treinamento = turma.treinamento
        turma_email_ctx = SimpleNamespace(
            treinamento=SimpleNamespace(
                nome=getattr(treinamento, "nome", ""),
                codigo=getattr(treinamento, "codigo", ""),
                carga_horaria=getattr(treinamento, "carga_horaria", None),
                tem_pratica=getattr(treinamento, "tem_pratica", False),
            ),
            data_inicio=turma.data_inicio,
            data_fim=turma.data_fim,
            horario=turma.horario,
            instrutor=turma.instrutor,
            local_realizacao=turma.local_realizacao,
            local_pratica=getattr(turma, "local_pratica", None),
            teoria_online=turma.teoria_online,
        )

        dados_log = {
            "id": turma.id,
            "treinamento_id": turma.treinamento_id,
            "treinamento_nome": (
                turma.treinamento.nome if turma.treinamento else None
            ),
            "data_inicio": turma.data_inicio.isoformat()
            if turma.data_inicio
            else None,
            "data_fim": turma.data_fim.isoformat() if turma.data_fim else None,
            "instrutor_id": turma.instrutor_id,
            "instrutor_nome": turma.instrutor.nome if turma.instrutor else None,
        }

        InscricaoTreinamento.query.filter_by(turma_id=turma_id).delete()
        db.session.delete(turma)
        db.session.commit()

        log_action(
            g.current_user.id,
            'delete',
            'TurmaTreinamento',
            dados_log["id"],
            dados_log,
        )

        if unique_recipients:
            try:
                send_treinamento_desmarcado_email(unique_recipients, turma_email_ctx)
            except Exception as exc:  # pragma: no cover - log apenas
                log.error(
                    "Erro ao enviar e-mail de treinamento desmarcado: %s", exc
                )

        return jsonify({"mensagem": "Turma removida com sucesso"})
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route("/treinamentos/turmas/<int:turma_id>/inscricoes", methods=["GET"])
@admin_required
def listar_inscricoes(turma_id):
    """Lista inscrições de uma turma."""
    turma = db.session.get(TurmaTreinamento, turma_id)
    if not turma:
        return jsonify({"erro": "Turma não encontrada"}), 404
    inscricoes = InscricaoTreinamento.query.filter_by(turma_id=turma_id).all()
    return jsonify([i.to_dict() for i in inscricoes])


@treinamento_bp.route(
    "/treinamentos/turmas/<int:turma_id>/inscricoes/admin", methods=["POST"]
)
@admin_required
def criar_inscricao_admin(turma_id):
    """Adiciona manualmente uma inscrição em uma turma."""
    turma = db.session.get(TurmaTreinamento, turma_id)
    if not turma:
        return jsonify({"erro": "Turma não encontrada"}), 404
    data = request.json or {}
    try:
        payload = InscricaoTreinamentoCreateSchema(**data)
    except ValidationError as e:
        return jsonify({"erro": e.errors()}), 400
    insc = InscricaoTreinamento(
        usuario_id=None,
        turma_id=turma_id,
        nome=payload.nome,
        email=payload.email,
        cpf=payload.cpf,
        data_nascimento=payload.data_nascimento,
        empresa=payload.empresa,
    )
    try:
        db.session.add(insc)
        db.session.commit()
        dados_log = {
            "id": insc.id,
            "nome_inscrito": insc.nome,
            "turma_id": turma.id,
            "nome_treinamento": turma.treinamento.nome if turma.treinamento else "N/A",
        }
        log_action(g.current_user.id, 'create', 'InscricaoTreinamento', insc.id, dados_log)
        return jsonify(insc.to_dict()), 201
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route(
    "/treinamentos/turmas/<int:turma_id>/inscricoes/export", methods=["GET"]
)
@admin_required
def exportar_inscricoes(turma_id):
    """Exporta inscrições de uma turma em CSV, XLSX ou PDF."""
    turma = db.session.get(TurmaTreinamento, turma_id)
    if not turma:
        return jsonify({"erro": "Turma não encontrada"}), 404

    formato = request.args.get("formato", "xlsx").lower()
    inscricoes = (
        InscricaoTreinamento.query.filter_by(turma_id=turma_id)
        .order_by(InscricaoTreinamento.nome)
        .all()
    )
    treinamento = turma.treinamento

    nome_arquivo_base = treinamento.nome.replace(" ", "_").lower()
    data_hoje = datetime.now().strftime("%Y-%m-%d")
    nome_arquivo_final = f"{nome_arquivo_base}_{data_hoje}"

    def format_date(dt):
        return dt.strftime("%d/%m/%Y") if dt else ""

    # --- LÓGICA PARA CSV ---
    if formato == "csv":
        si = StringIO()
        writer = csv.writer(si)
        headers = [
            "Nome",
            "E-mail",
            "CPF",
            "Empresa",
            "Presença Teoria",
            "Presença Prática",
            "Nota Teoria",
            "Nota Prática",
            "Status",
        ]
        writer.writerow(headers)
        for i in inscricoes:
            row = [
                i.nome,
                i.email,
                i.cpf,
                i.empresa,
                "Sim" if i.presenca_teoria else "Não",
                "Sim" if i.presenca_pratica else "Não",
                i.nota_teoria,
                i.nota_pratica,
                i.status_aprovacao,
            ]
            writer.writerow(row)
        output = make_response(si.getvalue())
        output.headers["Content-Disposition"] = (
            f"attachment; filename={nome_arquivo_final}.csv"
        )
        output.headers["Content-Type"] = "text/csv"
        return output

    # --- LÓGICA PARA XLSX ---
    if formato == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Lista de Presença"

        # --- Estilos ---
        cor_azul_senai_hex = "00539F"
        fill_azul = PatternFill(
            start_color=cor_azul_senai_hex,
            end_color=cor_azul_senai_hex,
            fill_type="solid",
        )
        font_white_bold = Font(color="FFFFFF", bold=True)
        font_bold = Font(bold=True)
        thin_border_side = Side(style="thin")
        thin_border = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side,
        )

        # --- Cabeçalho ---
        ws.merge_cells("A1:B2")
        ws.merge_cells("C1:K2")

        # Logo
        logo_path = _get_logo_path()
        try:
            if logo_path.exists():
                img = OpenpyxlImage(str(logo_path))
                img.anchor = "A1"
                img.height = 40
                ws.add_image(img)
            else:
                raise FileNotFoundError(str(logo_path))
        except FileNotFoundError:
            log.warning("Logo do SENAI não encontrado em %s", logo_path)
            logo_cell = ws["A1"]
            logo_cell.value = "SENAI"
            logo_cell.font = font_white_bold
            logo_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Título
        title_cell = ws["C1"]
        title_cell.value = "Lista de Presença"
        title_cell.font = Font(color="FFFFFF", bold=True, size=20)
        title_cell.fill = fill_azul
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Cor de fundo na célula da logo
        for row in ws["A1:B2"]:
            for cell in row:
                cell.fill = fill_azul

        # --- Tabela de Dados do Treinamento ---
        row_idx = 4
        dados_treinamento = {
            "Unidade:": "SENAI - Conceição do Mato Dentro",
            "Nome Treinamento:": treinamento.nome,
            "Instituição:": "SENAI",
            "Local de Realização:": turma.local_realizacao or "N/D",
            "Instrutor(es):": turma.instrutor.nome if turma.instrutor else "N/D",
            "CONTEÚDO PROGRAMÁTICO:": (treinamento.conteudo_programatico or "").replace(
                "\n", "\n"
            ),
        }

        dados_treinamento_lado_direito = {
            "Período:": f"{format_date(turma.data_inicio)} a {format_date(turma.data_fim)}",
            "Duração:": f"{treinamento.carga_horaria or 'N/D'} horas",
            "Horário:": turma.horario or "N/D",
        }

        # Aplicar borda em toda a área da tabela de dados
        for col in "ABCDEFGHIJ":
            for row in range(row_idx, row_idx + 6):
                ws[f"{col}{row}"].border = thin_border

        # Preencher dados
        for i, (label, value) in enumerate(dados_treinamento.items()):
            current_row = row_idx + i
            label_cell = ws[f"A{current_row}"]
            label_cell.value = label
            label_cell.fill = fill_azul
            label_cell.font = font_white_bold
            label_cell.alignment = Alignment(vertical="top")

            value_cell = ws[f"B{current_row}"]
            value_cell.value = value
            value_cell.alignment = Alignment(vertical="top", wrap_text=True)

            if label in ["Instituição:", "Local de Realização:", "Instrutor(es):"]:
                ws.merge_cells(f"B{current_row}:F{current_row}")
            else:
                ws.merge_cells(f"B{current_row}:I{current_row}")

        # Lado direito
        ws["G6"].value = "Período:"
        ws["H6"].value = dados_treinamento_lado_direito["Período:"]
        ws["G7"].value = "Duração:"
        ws["H7"].value = dados_treinamento_lado_direito["Duração:"]
        ws["G8"].value = "Horário:"
        ws["H8"].value = dados_treinamento_lado_direito["Horário:"]

        ws.merge_cells("B6:F6")
        ws.merge_cells("B7:F7")
        ws.merge_cells("B8:F8")
        ws.merge_cells("H6:I6")
        ws.merge_cells("H7:I7")
        ws.merge_cells("H8:I8")

        for cell_label, cell_value in [("G6", "H6"), ("G7", "H7"), ("G8", "H8")]:
            ws[cell_label].fill = fill_azul
            ws[cell_label].font = font_white_bold
            ws[cell_label].alignment = Alignment(vertical="top")
            ws[cell_value].alignment = Alignment(vertical="top")

        row_idx += 7

        # --- Tabela de Participantes ---
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        info_cell = ws[f"A{row_idx}"]
        info_cell.value = "Informações dos participantes"
        info_cell.fill = fill_azul
        info_cell.font = font_white_bold
        info_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"G{row_idx}:K{row_idx}")
        rubrica_cell = ws[f"G{row_idx}"]
        rubrica_cell.value = "Rubrica do participante conforme data de participação"
        rubrica_cell.fill = fill_azul
        rubrica_cell.font = font_white_bold
        rubrica_cell.alignment = Alignment(horizontal="center", vertical="center")

        row_idx += 1
        headers = [
            "Nº",
            "CPF",
            "Data de Nascimento",
            "Nome do Participante",
            "E-mail",
            "Empresa",
            "TEORIA",
            "NOTA DA\nTEORIA",
            "PRÁTICA",
            "NOTA DA\nPRÁTICA",
            "APROVADO /\nREPROVADO",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.fill = fill_azul
            cell.font = font_white_bold
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        row_idx += 1
        for i, inscricao in enumerate(inscricoes, 1):
            ws.cell(row=row_idx, column=1, value=i)
            ws.cell(row=row_idx, column=2, value=inscricao.cpf)
            ws.cell(
                row=row_idx,
                column=3,
                value=(
                    inscricao.data_nascimento.strftime("%d/%m/%Y")
                    if inscricao.data_nascimento
                    else ""
                ),
            )
            ws.cell(row=row_idx, column=4, value=inscricao.nome)
            ws.cell(row=row_idx, column=5, value=inscricao.email)
            ws.cell(row=row_idx, column=6, value=inscricao.empresa)
            row_idx += 1

        # Borda na tabela de participantes
        for col in "ABCDEFGHIJK":
            for row in range(row_idx - len(inscricoes) - 2, row_idx):
                ws[f"{col}{row}"].border = thin_border
                ws[f"{col}{row}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )

        # --- Observações e Assinatura ---
        row_idx += 1
        ws.merge_cells(f"A{row_idx}:K{row_idx+2}")
        obs_cell = ws[f"A{row_idx}"]
        obs_cell.value = "Observações:"
        obs_cell.font = font_bold
        obs_cell.alignment = Alignment(horizontal="left", vertical="top")
        obs_cell.border = thin_border

        row_idx += 4
        ws.merge_cells(f"A{row_idx}:K{row_idx+1}")
        ass_cell = ws[f"A{row_idx}"]
        ass_cell.value = "Assinatura do(s) instrutor(es) / Responsável (eis):"
        ass_cell.font = font_bold
        ass_cell.alignment = Alignment(horizontal="left", vertical="top")
        ass_cell.border = Border(bottom=thin_border_side)

        # --- Ajustar tamanho das colunas e linhas ---
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 10
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 15
        ws.row_dimensions[9].height = 40

        # --- Salvar em buffer ---
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"{nome_arquivo_final}.xlsx",
        )

    # --- LÓGICA PARA PDF ---
    if formato == "pdf":
        buffer = BytesIO()
        # Redução das margens e ajuste dos tamanhos de fonte
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=18,
            leftMargin=18,
            topMargin=18,
            bottomMargin=18,
        )
        elements = []
        styles = getSampleStyleSheet()

        cor_azul_senai_rgb = colors.Color(
            red=(0 / 255), green=(83 / 255), blue=(159 / 255)
        )

        style_normal = ParagraphStyle(
            name="Normal", fontSize=6.5, leading=7.5
        )  # Fonte reduzida
        style_bold_white = ParagraphStyle(
            name="BoldWhite",
            parent=style_normal,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )
        style_h1_centralizado = ParagraphStyle(
            name="h1_centralizado",
            parent=styles["h1"],
            alignment=1,
            textColor=colors.white,
            fontSize=14,  # Fonte reduzida
        )

        try:
            logo_path = _get_logo_path()
            if logo_path.exists():
                logo = ReportlabImage(
                    str(logo_path), width=1.2 * inch, height=0.4 * inch  # Logo menor
                )
                logo.hAlign = "CENTER"
            else:
                raise FileNotFoundError(str(logo_path))
        except Exception as exc:
            log.warning("Falha ao carregar o logo do SENAI: %s", exc)
            logo = Paragraph("<b>SENAI</b>", style_normal)

        titulo = Paragraph("<b>Lista de Presença</b>", style_h1_centralizado)

        header_logo_width = 1.45 * inch
        header_table = Table(
            [[logo, titulo]],
            colWidths=[header_logo_width, doc.width - header_logo_width],
        )
        header_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BACKGROUND", (0, 0), (-1, -1), cor_azul_senai_rgb),
                ]
            )
        )
        elements.append(header_table)
        elements.append(Spacer(1, 0.05 * inch))  # Espaçamento reduzido

        # Tabela de dados do treinamento
        dados_treinamento = [
            [
                Paragraph("<b>Unidade:</b>", style_bold_white),
                Paragraph("SENAI - Conceição do Mato Dentro", style_normal),
                None,
                None,
            ],
            [
                Paragraph("<b>Nome Treinamento:</b>", style_bold_white),
                Paragraph(treinamento.nome, style_normal),
                None,
                None,
            ],
            [
                Paragraph("<b>Instituição:</b>", style_bold_white),
                Paragraph("SENAI", style_normal),
                Paragraph("<b>Período:</b>", style_bold_white),
                Paragraph(
                    f"{format_date(turma.data_inicio)} a {format_date(turma.data_fim)}",
                    style_normal,
                ),
            ],
            [
                Paragraph("<b>Local de Realização:</b>", style_bold_white),
                Paragraph(turma.local_realizacao or "N/D", style_normal),
                Paragraph("<b>Duração:</b>", style_bold_white),
                Paragraph(f"{treinamento.carga_horaria or 'N/D'} horas", style_normal),
            ],
            [
                Paragraph("<b>Instrutor(es):</b>", style_bold_white),
                Paragraph(
                    turma.instrutor.nome if turma.instrutor else "N/D", style_normal
                ),
                Paragraph("<b>Horário:</b>", style_bold_white),
                Paragraph(turma.horario or "N/D", style_normal),
            ],
            [
                Paragraph("<b>CONTEÚDO PROGRAMÁTICO:</b>", style_bold_white),
                Paragraph(
                    (treinamento.conteudo_programatico or "").replace("\n", "<br/>"),
                    style_normal,
                ),
                None,
                None,
            ],
        ]

        col1 = 1.25 * inch
        col3 = 1.0 * inch
        col4 = 1.45 * inch
        col2 = doc.width - (col1 + col3 + col4)
        if col2 < 2.0 * inch:
            col2 = 2.0 * inch
            col4 = doc.width - (col1 + col2 + col3)
        if col4 < 1.2 * inch:
            col4 = 1.2 * inch
            col2 = doc.width - (col1 + col3 + col4)
        tabela_dados = Table(
            dados_treinamento,
            colWidths=[col1, col2, col3, col4],
        )
        tabela_dados.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BOX", (0, 0), (-1, -1), 1, colors.black),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
                    ("SPAN", (1, 0), (-1, 0)),
                    ("SPAN", (1, 1), (-1, 1)),
                    ("SPAN", (1, 5), (-1, 5)),
                    ("BACKGROUND", (0, 0), (0, -1), cor_azul_senai_rgb),
                    ("BACKGROUND", (2, 2), (2, 4), cor_azul_senai_rgb),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                    ("TEXTCOLOR", (2, 2), (2, 4), colors.white),
                ]
            )
        )
        elements.append(tabela_dados)
        elements.append(Spacer(1, 0.1 * inch))  # Espaçamento reduzido

        # Estilo para cabeçalhos da tabela de participantes
        style_header_participantes = ParagraphStyle(
            name="HeaderParticipantes",
            fontSize=5.8,  # Fonte reduzida
            leading=6.6,
            alignment=1,
            fontName="Helvetica-Bold",
            textColor=colors.white,
        )

        # Cabeçalhos com quebra de linha
        tabela_header = [
            "Nº",
            "Nome do Participante",
            "Empresa",
            "TEORIA",
            Paragraph("NOTA DA<br/>TEORIA", style_header_participantes),
            "PRÁTICA",
            Paragraph("NOTA DA<br/>PRÁTICA", style_header_participantes),
            Paragraph("APROVADO /<br/>REPROVADO", style_header_participantes),
        ]

        cabecalhos_agrupados = [
            [
                Paragraph("<b>Informações dos participantes</b>", style_bold_white),
                None,
                None,
                Paragraph(
                    "<b>Rubrica do participante conforme data de participação</b>",
                    style_bold_white,
                ),
                None,
                None,
                None,
                None,
            ],
            [
                (
                    Paragraph(f"<b>{h}</b>", style_header_participantes)
                    if isinstance(h, str)
                    else h
                )
                for h in tabela_header
            ],
        ]

        dados_alunos = []
        for idx, i in enumerate(inscricoes, 1):
            dados_alunos.append(
                [
                    str(idx),
                    Paragraph(i.nome, style_normal),
                    i.empresa or "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

        col_widths = [
            0.05 * doc.width,
            0.34 * doc.width,
            0.2 * doc.width,
            0.08 * doc.width,
            0.08 * doc.width,
            0.08 * doc.width,
            0.08 * doc.width,
            0.09 * doc.width,
        ]

        # Define a altura das linhas para economizar espaço
        num_participantes = len(dados_alunos)
        alturas_linhas = [
            0.28 * inch,
            0.26 * inch,
        ] + [0.18 * inch] * num_participantes  # Alturas reduzidas

        tabela_alunos = Table(
            cabecalhos_agrupados + dados_alunos,
            colWidths=col_widths,
            rowHeights=alturas_linhas,
        )

        tabela_alunos.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 1), cor_azul_senai_rgb),
                    ("TEXTCOLOR", (0, 0), (-1, 1), colors.white),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BOX", (0, 0), (-1, -1), 1, colors.black),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
                    ("SPAN", (0, 0), (2, 0)),
                    ("SPAN", (3, 0), (-1, 0)),
                ]
            )
        )

        # Observações e Assinatura
        obs_table = Table(
            [[Paragraph("<b>Observações:</b>", style_normal)], [""]],
            colWidths=[doc.width],
            rowHeights=[0.14 * inch, 0.35 * inch],  # Alturas reduzidas
        )
        obs_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )

        ass_table = Table(
            [
                [
                    Paragraph(
                        "<b>Assinatura do(s) instrutor(es) / Responsável (eis):</b>",
                        style_normal,
                    )
                ],
                [""],
            ],
            colWidths=[doc.width],
            rowHeights=[0.14 * inch, 0.22 * inch],  # Alturas reduzidas
        )
        ass_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 1, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LINEBELOW", (0, 1), (0, 1), 1, colors.black),
                ]
            )
        )

        # Agrupa os últimos elementos para evitar quebra de página
        conteudo_final = KeepTogether(
            [
                tabela_alunos,
                Spacer(1, 0.1 * inch),  # Espaçamento reduzido
                obs_table,
                Spacer(1, 0.1 * inch),  # Espaçamento reduzido
                ass_table,
            ]
        )
        elements.append(conteudo_final)

        doc.build(elements)

        buffer.seek(0)
        return send_file(
            buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"{nome_arquivo_final}.pdf",
        )

    return jsonify({"erro": "Formato inválido"}), 400


@treinamento_bp.route("/treinamentos/turmas/<int:turma_id>", methods=["GET"])
@admin_required
def obter_turma_treinamento(turma_id):
    """Obtém detalhes de uma turma de treinamento."""
    turma = db.session.get(TurmaTreinamento, turma_id)
    if not turma:
        return jsonify({"erro": "Turma não encontrada"}), 404

    dados_turma = turma.to_dict()
    dados_turma["treinamento"] = (
        turma.treinamento.to_dict() if turma.treinamento else None
    )
    dados_turma["instrutor"] = turma.instrutor.to_dict() if turma.instrutor else None

    return jsonify(dados_turma)


@treinamento_bp.route(
    "/treinamentos/inscricoes/<int:inscricao_id>/avaliar", methods=["PUT"]
)
@admin_required
def avaliar_inscricao(inscricao_id):
    """Atualiza as notas e o status de aprovação de uma inscrição."""
    inscricao = db.session.get(InscricaoTreinamento, inscricao_id)
    if not inscricao:
        return jsonify({"erro": "Inscrição não encontrada"}), 404

    data = request.json
    if not data:
        return jsonify({"erro": "Dados não fornecidos"}), 400

    try:
        nota_teoria = data.get("nota_teoria")
        nota_pratica = data.get("nota_pratica")

        inscricao.nota_teoria = (
            float(nota_teoria) if nota_teoria not in [None, ""] else None
        )
        inscricao.nota_pratica = (
            float(nota_pratica) if nota_pratica not in [None, ""] else None
        )
        inscricao.status_aprovacao = data.get("status_aprovacao")

        # Campos de presença
        inscricao.presenca_teoria = data.get("presenca_teoria", False)
        inscricao.presenca_pratica = data.get("presenca_pratica", False)

        db.session.commit()
        return jsonify(inscricao.to_dict())

    except (ValueError, TypeError):
        db.session.rollback()
        return jsonify({"erro": "Valores de nota inválidos. Devem ser números."}), 400
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.route("/treinamentos/inscricoes/<int:inscricao_id>", methods=["DELETE"])
@admin_required
def remover_inscricao(inscricao_id):
    """Remove uma inscrição de uma turma."""
    inscricao = db.session.get(InscricaoTreinamento, inscricao_id)
    if not inscricao:
        return jsonify({"erro": "Inscrição não encontrada"}), 404

    try:
        dados_log = {
            "id": inscricao.id,
            "nome_inscrito": inscricao.nome,
            "turma_id": inscricao.turma_id,
            "nome_treinamento": inscricao.turma.treinamento.nome if inscricao.turma and inscricao.turma.treinamento else "N/A",
        }
        db.session.delete(inscricao)
        db.session.commit()
        log_action(
            g.current_user.id,
            'delete',
            'InscricaoTreinamento',
            inscricao.id,
            dados_log,
        )
        return jsonify({"mensagem": "Inscrição removida com sucesso"}), 200
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)


@treinamento_bp.post("/inscricoes/<int:inscricao_id>/convocar")
@admin_required
def convocar_inscrito(inscricao_id: int):
    """Envia e-mail de convocação para o inscrito."""
    insc = db.session.get(InscricaoTreinamento, inscricao_id)
    if not insc or not insc.email:
        return jsonify({"erro": "Inscrição não encontrada ou sem e-mail"}), 404

    turma = db.session.get(TurmaTreinamento, insc.turma_id)
    if not turma:
        return jsonify({"erro": "Turma não encontrada"}), 404

    treino = turma.treinamento
    if not treino:
        return jsonify({"erro": "Treinamento não encontrado"}), 404

    enviar_convocacao(insc, turma, send_email_fn=send_email)

    if hasattr(insc, "convocado_em"):
        insc.convocado_em = datetime.utcnow()
        db.session.commit()

    return jsonify({"ok": True})


@treinamento_bp.route(
    "/treinamentos/<int:turma_id>/inscricoes/externo", methods=["POST"]
)
@login_required
def create_inscricao_treinamento_externo(turma_id):
    """Cria uma nova inscrição para participante externo."""
    data = request.get_json()

    payload = InscricaoTreinamentoCreateSchema(**data)

    turma = TurmaTreinamento.query.get_or_404(turma_id)

    nova_inscricao = InscricaoTreinamento(
        turma_id=turma.id,
        usuario_id=None,
        nome=payload.nome,
        email=payload.email,
        cpf=payload.cpf,
        data_nascimento=payload.data_nascimento,
        empresa=payload.empresa,
    )

    db.session.add(nova_inscricao)
    db.session.commit()

    return jsonify(nova_inscricao.to_dict()), 201


@treinamento_bp.route("/treinamentos/logs", methods=["GET"])
@admin_required
def listar_logs_treinamentos():
    """Lista os logs de auditoria relacionados a treinamentos e inscrições."""
    try:
        logs = (
            db.session.query(AuditLog, User.nome)
            .join(User, User.id == AuditLog.user_id)
            .filter(
                AuditLog.entity.in_([
                    'Treinamento',
                    'TurmaTreinamento',
                    'InscricaoTreinamento',
                ])
            )
            .order_by(AuditLog.timestamp.desc())
            .all()
        )

        resultado = []
        for log, nome_usuario in logs:
            detalhes = log.details or {}
            info = ""

            if log.entity == 'InscricaoTreinamento' and log.action == 'create':
                nome_treinamento = detalhes.get('nome_treinamento', 'N/A')
                nome_inscrito = detalhes.get('nome_inscrito', 'N/A')
                turma_id = detalhes.get('turma_id', 'N/A')
                info = (
                    f"Inscrito '{nome_inscrito}' adicionado ao treinamento "
                    f"'{nome_treinamento}' (Turma ID: {turma_id})"
                )
            elif log.entity == 'InscricaoTreinamento' and log.action == 'delete':
                nome_treinamento = detalhes.get('nome_treinamento', 'N/A')
                nome_inscrito = detalhes.get('nome_inscrito', 'N/A')
                turma_id = detalhes.get('turma_id', 'N/A')
                info = (
                    f"Inscrito '{nome_inscrito}' removido do treinamento "
                    f"'{nome_treinamento}' (Turma ID: {turma_id})"
                )
            elif log.entity == 'InscricaoTreinamento':
                info = f"Inscrição '{detalhes.get('nome', '')}' (ID: {log.entity_id})"
            elif log.entity == 'TurmaTreinamento':
                info = (
                    f"Turma do treinamento '{detalhes.get('treinamento_nome', 'N/A')}' "
                    f"(ID: {log.entity_id})"
                )
            else:
                info = f"Treinamento '{detalhes.get('nome', '')}' (ID: {log.entity_id})"

            resultado.append({
                "id": log.id,
                "timestamp": log.timestamp.isoformat(),
                "usuario": nome_usuario,
                "acao": log.action,
                "info": info,
            })

        return jsonify(resultado)
    except SQLAlchemyError as e:
        db.session.rollback()
        return handle_internal_error(e)
